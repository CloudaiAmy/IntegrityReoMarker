"""Extract and normalize column reinforcement data from Excel spreadsheets."""

import re
from typing import Any, Optional

from openpyxl import load_workbook

from models import ColumnReinforcement, ReinforcementRegistry

# Output key -> acceptable header patterns (matched after normalization)
COLUMN_SPECS: list[tuple[str, list[str]]] = [
    ("mark", ["reference", "ref"]),
    (
        "dimension_x",
        [
            "column dimension x mm",
            "column dimension x (mm)",
            "column dimension x",
            "col dimension x",
            "x mm",  # only when grouped under column dimension
        ],
    ),
    (
        "dimension_y",
        [
            "column dimension y mm",
            "column dimension y (mm)",
            "column dimension y",
            "col dimension y",
        ],
    ),
    ("edge_condition", ["edge condition"]),
    (
        "slab_reaction_n",
        ["slab reaction n kn", "slab reaction kn", "slab reaction n"],
    ),
    ("fc_slab", ["f'c slab mpa", "fc slab mpa", "f'c slab", "fc slab"]),
    ("fsy", ["fsy mpa", "fsy"]),
    ("bar_diameter", ["bar diameter", "bar dia"]),
    ("as_required", ["as mm2", "as (mm2)"]),
    ("as_total", ["total mm2", "total (mm2)"]),
    ("as_min", ["as.min total mm2", "as min total mm2", "as.min", "as min"]),
    (
        "bars_required_x",
        ["# bars required x", "bars required x", "no bars required x"],
    ),
    (
        "bars_required_y",
        ["# bars required y", "bars required y", "no bars required y"],
    ),
    (
        "bar_length_x",
        ["bar length x mm", "bar length x (mm)", "bar length x"],
    ),
    (
        "bar_length_y",
        ["bar length y mm", "bar length y (mm)", "bar length y"],
    ),
]

STRING_FIELDS = frozenset({"mark", "edge_condition"})
INTEGER_FIELDS = frozenset(
    {
        "bars_required_x",
        "bars_required_y",
        "bar_diameter",
        "dimension_x",
        "dimension_y",
        "bar_length_x",
        "bar_length_y",
    }
)

# Composite headers must include group context for ambiguous short names like "x mm"
_GROUP_REQUIRED_PATTERNS = frozenset({"x mm", "y mm"})

_COLUMN_REFERENCE_RE = re.compile(r"^[A-Za-z]+\d+$")


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("’", "'").replace("′", "'")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[()（）]", "", text)
    return text.strip()


def _find_reference_header_row(rows: list[tuple[Any, ...]]) -> int:
    """Return the row index that contains the Reference column header."""
    for index, row in enumerate(rows):
        for cell in row:
            if _normalize_header(cell) == "reference":
                return index
    raise ValueError(
        "Could not find a Reference column. Expected a header like 'Reference'."
    )


def _row_has_content(row: tuple[Any, ...]) -> bool:
    return any(cell is not None and str(cell).strip() for cell in row)


def _resolve_group_label(
    index: int,
    group_row: Optional[tuple[Any, ...]],
    header_row: tuple[Any, ...],
) -> str:
    """Resolve group label for a column, including merged multi-column group headers."""
    if group_row and index < len(group_row):
        cell = group_row[index]
        if cell is not None and str(cell).strip():
            return str(cell).strip()

    sub = _normalize_header(header_row[index] if index < len(header_row) else "")
    if not sub:
        return ""

  # Look left for a group header that owns this sub-column (report-style layouts).
    for lookback in range(index - 1, max(index - 3, -1), -1):
        if not group_row or lookback >= len(group_row):
            continue
        group_cell = group_row[lookback]
        if group_cell is None or not str(group_cell).strip():
            continue
        group = str(group_cell).strip()
        group_norm = _normalize_header(group)

        if sub in ("x mm", "y mm") and "column dimension" in group_norm:
            return group
        if sub in ("x", "y") and "bars required" in group_norm:
            return group
        if sub in ("x mm", "y mm") and "bar length" in group_norm:
            return group
        break

    return ""


def _build_composite_headers(
    group_row: Optional[tuple[Any, ...]],
    header_row: tuple[Any, ...],
) -> list[str]:
    """
    Combine optional group header row (e.g. COLUMN DIMENSION) with sub-headers
    (e.g. X (mm)) for report-style spreadsheets like BAE46300.
    """
    length = len(header_row)
    if group_row:
        length = max(length, len(group_row))

    headers: list[str] = []

    for index in range(length):
        header_cell = header_row[index] if index < len(header_row) else None
        sub = str(header_cell).strip() if header_cell is not None else ""
        group = _resolve_group_label(index, group_row, header_row)

        if sub and group:
            headers.append(f"{group} {sub}")
        elif sub:
            headers.append(sub)
        elif group:
            headers.append(group)
        else:
            headers.append("")

    return headers


def normalize_headers(headers: list[Any]) -> dict[str, int]:
    """
    Map output keys to column indices from composite or single-row headers.

    Raises ValueError if the Reference column cannot be found.
    """
    normalized = [_normalize_header(h) for h in headers]
    column_map: dict[str, int] = {}
    used_indices: set[int] = set()

    for output_key, patterns in COLUMN_SPECS:
        patterns_sorted = sorted(patterns, key=len, reverse=True)
        for index, header in enumerate(normalized):
            if index in used_indices or not header:
                continue
            for pattern in patterns_sorted:
                pattern_norm = _normalize_header(pattern)
                if pattern_norm not in _GROUP_REQUIRED_PATTERNS:
                    if header == pattern_norm or pattern_norm in header:
                        column_map[output_key] = index
                        used_indices.add(index)
                        break
                elif header == pattern_norm and "column dimension" in header:
                    column_map[output_key] = index
                    used_indices.add(index)
                    break
            if output_key in column_map:
                break

    if "mark" not in column_map:
        raise ValueError(
            "Could not find a Reference column. Expected a header like 'Reference'."
        )

    return column_map


def _parse_number(value: Any) -> Optional[int | float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        return int(value) if value == int(value) else value

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(",", "")
    match = re.search(r"-?\d+\.?\d*", text)
    if not match:
        return None

    number = float(match.group())
    return int(number) if number == int(number) else number


def _clean_value(output_key: str, value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, str) and not value.strip():
        return None

    if output_key in STRING_FIELDS:
        text = str(value).strip()
        return text or None

    number = _parse_number(value)
    if number is None:
        return None

    if output_key in INTEGER_FIELDS and isinstance(number, float):
        return int(number)
    return number


def _looks_like_column_reference(text: str) -> bool:
    """True for column IDs like CC01, C1, C12 (not lookup-table labels)."""
    return bool(_COLUMN_REFERENCE_RE.match(text.strip()))


def extract_row(row: tuple[Any, ...], column_map: dict[str, int]) -> Optional[dict[str, Any]]:
    """Extract a normalized row dict. Returns None if Reference/mark is empty."""
    record: dict[str, Any] = {}

    for output_key, index in column_map.items():
        if index >= len(row):
            record[output_key] = None
            continue
        record[output_key] = _clean_value(output_key, row[index])

    mark = record.get("mark")
    if mark is None:
        return None

    mark_text = str(mark).strip()
    if not mark_text:
        return None

    record["mark"] = mark_text
    return record


def build_registry(path: str) -> ReinforcementRegistry:
    """
    Load spreadsheet into an in-memory ReinforcementRegistry.

    Duplicate references keep the last row.
    Supports report-style layouts with title rows and two-row grouped headers.
    """
    registry = ReinforcementRegistry(source_path=path)
    seen_marks: dict[str, str] = {}

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise ValueError("Spreadsheet is empty.")

        header_index = _find_reference_header_row(all_rows)
        group_row: Optional[tuple[Any, ...]] = None
        if header_index > 0 and _row_has_content(all_rows[header_index - 1]):
            group_row = all_rows[header_index - 1]

        composite_headers = _build_composite_headers(group_row, all_rows[header_index])
        column_map = normalize_headers(composite_headers)
        mark_column = column_map["mark"]

        for row_number, row in enumerate(
            all_rows[header_index + 1 :],
            start=header_index + 2,
        ):
            if row is None:
                continue

            mark_value = row[mark_column] if mark_column < len(row) else None
            if mark_value is None or not str(mark_value).strip():
                continue
            if not _looks_like_column_reference(str(mark_value)):
                continue

            record = extract_row(tuple(row), column_map)
            if record is None:
                continue

            column = ColumnReinforcement.from_record(record)
            mark = column.reference
            mark_key = mark.upper()

            if mark_key in seen_marks:
                canonical = seen_marks[mark_key]
                if canonical not in registry.duplicate_references:
                    registry.duplicate_references.append(canonical)
                registry.warnings.append(
                    f"Duplicate Reference '{mark}' on row {row_number}; keeping last."
                )
            else:
                seen_marks[mark_key] = mark

            registry.columns[mark] = column
    finally:
        wb.close()

    if not registry.columns:
        registry.warnings.append("No data rows with a Reference value were found.")

    return registry
